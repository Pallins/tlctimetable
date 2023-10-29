import openpyxl
import xlsxwriter
import aspose.pdf as ap
from os.path import exists

def PDFtoXLSX(file_in):
    file_out=file_in.replace("pdf", "xlsx")
    document=ap.Document(file_in)
    save_options=ap.ExcelSaveOptions()
    save_options.format=ap.ExcelSaveOptions.ExcelFormat.XLSX
    document.save(file_out, save_options)

#sorts coordinates of cells like A1, B1 ecc.
def letter1(e):
    return e[0:1]

#sorts coordinates of cells like AA, AB ecc.
def letter2(e):
    return e[1:2]


def orario(path, classe, p=0, dictionary={}):
    file_xlsx=path.replace("pdf", "xlsx")

    if exists(file_xlsx)==False:
        PDFtoXLSX(path)

    workbook_i=openpyxl.load_workbook(file_xlsx)
    foglio=workbook_i.active
    ########################
    #creo un file xlsx per l'orario e definisco lo stile
    if p==1:
        workbook_f=xlsxwriter.Workbook(f"C:\\Users\\teopa\\Downloads\\orarioM_{classe}.xlsx")
    else:
        workbook_f=xlsxwriter.Workbook(f"C:\\Users\\teopa\\Downloads\\orario_{classe}.xlsx")

    worksheet_f=workbook_f.add_worksheet()
    format_day=workbook_f.add_format()
    format_day.set_font_size(32)
    format_day.set_align("center")
    format_day.set_font_color("#2900d6")
    format_day.set_bg_color("#f3fa9c")
    format_day.set_bold()
    format_day.set_border(2)
    format_day.set_border_color("black")
    format_prof=workbook_f.add_format()
    format_prof.set_font_size(22)
    format_prof.set_align("center")
    format_prof.set_bg_color("#f3fa9c")
    format_prof.set_border(2)
    format_prof.set_border_color("black")
    ########################
    lista1=[]
    lista2=[]
    lista3=[]
    giorni=["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
    var_giorno=""
    timetable=""
    x=0
    y=0
    widths=[]
    width=0

    for row_d in foglio:
        for column_d in range(1, foglio.max_column):
            if "lun" in str(row_d[column_d].value).lower():
                pos_giorni=row_d[column_d].row
                break


    for row in foglio:
        for i in range(1, foglio.max_column):
            if classe in str(row[i].value):
                coordinate=row[i].coordinate
                lista1.append(coordinate)

    
    for h in range(len(lista1)):
        if lista1[h][0:2].isalpha()==False:
            lista2.append(lista1[h])
        else:
            lista3.append(lista1[h])
            
            
    
        
    lista2.sort(key=letter1)
    lista3.sort(key=letter2)

    lista=lista2+lista3
    


    for j in range(len(lista)):  

        colonna=""
        for k in lista[j]:
            if k.isdigit()==True:
                colonna=colonna+str(k)
        
        giornox=lista[j].replace(colonna, "")
        giorno_coordinata=giornox+str(pos_giorni)
        giorno=foglio[giorno_coordinata].value

        if giorno!=var_giorno:
            y=0
            var_giorno=giorno
            worksheet_f.write(y, x, giorni[0].upper(), format_day)
            timetable+=f"\n{giorni[0]}\n\n"
            giorni.remove(giorni[0])
            x+=1
            if width!=0:
                widths.append(width)
                width=0

        y+=1
        coordinata_prof="A"+colonna
        if p==1:
            try:
                prof=dictionary[str(foglio[coordinata_prof].value).lower()].upper()
            except:
                prof=str(foglio[coordinata_prof].value).upper()
        else:
            prof=str(foglio[coordinata_prof].value).upper()

        worksheet_f.write(y, (x-1), prof, format_prof)
        timetable+=f"{prof}\n"
        if len(prof)>width:
            width=len(prof)

    widths.append(width)
    
    giorni=["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
    for column in range(x):
        len_d=len(giorni[column])*4
        len_p=widths[column]*2.6
        if len_p>=len_d:
            worksheet_f.set_column(column, column, width=int(len_p))
        else:
            worksheet_f.set_column(column, column, width=int(len_d))

    workbook_f.close()
    print("Done")
    return timetable


subject={"donnini antonella":"italiano", "vilmercati francesca":"matematica", "larocca francesco": "scienze", "di giannantonio anto":"inglese", \
"rossitto carmela":"arte", "gentili barbara":"storia", "silvestrini davide":"religione", "zanoboni marzia": "ed. fisica", "garofalo simona":"fisica"}


orario("C:\\Users\\teopa\\Downloads\\orario completo 9 -14 Ottobre prot.pdf", "4A", 1, subject)